import openpyxl
import os
import yt_dlp
import pandas as pd
from googleapiclient.discovery import build
from pytubefix import YouTube


def get_youtube_video_links(api_key, keyword, max_results):
    youtube = build('youtube', 'v3', developerKey=api_key)

    # Load nextPageToken if it exists
    next_page_token = None
    if os.path.exists('nextpage.txt'):
        with open('nextpage.txt', 'r') as file:
            next_page_token = file.read().strip()

    video_links = []

    while True:
        search_request = youtube.search().list(
            part='snippet',
            q=keyword,
            type='video',
            maxResults=max_results,
            pageToken=next_page_token
        )

        search_response = search_request.execute()

        for item in search_response['items']:
            video_id = item['id']['videoId']
            video_title = item['snippet']['title']
            video_url = f"https://www.youtube.com/watch?v={video_id}"
            video_links.append({
                'Title': clean_title(video_title),
                'Link': video_url
            })

        # Check for the next page token and save it to a file
        next_page_token = search_response.get('nextPageToken', None)
        if next_page_token:
            with open('nextpage.txt', 'w') as file:
                file.write(next_page_token)
        else:
            # If there's no next page, remove the token file
            if os.path.exists('nextpage.txt'):
                os.remove('nextpage.txt')
            break

    return video_links


def save_video_links_to_excel(api_key,keyword,max_results):

    filename=f'./output_files/{keyword}/processing/youtube_video_links.xlsx'

    video_links= get_youtube_video_links(api_key, keyword, max_results)
    # Load existing workbook or create a new one
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title='Video Links')

        # Create headers for the new sheet
        headers = ['Title', 'Link']
        sheet.append(headers)

    sheet = workbook['Video Links']

    # Add video links
    for video in video_links:
        sheet.append([video['Title'], video['Link']])

    # Save the workbook
    workbook.save(filename)


# def clean_title(title):
#     return "".join([c for c in title if c.isalpha() or c.isdigit() or c == " "]).rstrip()[:-3]
# df = pd.read_excel('./output_files/CMO/CMO_video_analysis_results.xlsx', sheet_name='Sheet1', engine='openpyxl')
# df['Title'] = df['Title'].apply(clean_title)
# df.to_excel('./output_files/CMO/CMO_video_analysis_results_v2.xlsx', index=False)

df1=pd.read_excel('./output_files/CMO/CMO_video_analysis_results_v2.xlsx', sheet_name='Sheet1', engine='openpyxl')
df2=pd.read_excel('./output_files/CMO/processing/Processed_Video_Links.xlsx', sheet_name='Sheet1', engine='openpyxl')
merged_data=pd.merge(df2,df1,how='inner',on='Title')
merged_data.to_excel('./output_files/CMO/CMO_videos_data.xlsx',index=False)
print("Merge complete. Saved as 'CMO_videos_data.xlsx'.")


