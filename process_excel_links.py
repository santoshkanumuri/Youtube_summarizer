import pandas as pd
import re
from openpyxl import load_workbook
import yt_dlp
from pytubefix import YouTube
import os
from summarizer import analyze_folder


def get_links_from_csv(file_path,name):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb['Video links']

    # Function to extract YouTube links from a hyperlink or plain text
    def extract_youtube_links_from_rows(cell):
        if cell is None:
            return []
        # Extract hyperlink if present
        link = None
        if cell.hyperlink:
            link = cell.hyperlink.target
        else:
            link = cell.value

        # Regex pattern to match YouTube links
        youtube_pattern = r'(https?://(?:www\.)?youtube\.com/watch\?v=[\w-]+|https?://youtu\.be/[\w-]+)'
        return re.findall(youtube_pattern, str(link))

    # Create a new DataFrame to store the processed data
    processed_data = []

    # Iterate through each row in the sheet
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row is the header
        col1 = row[0].value  # First column
        col2 = row[1].value  # Second column
        col3 = row[2].value.replace(" ", "_") if row[
            2].value else None  # Replace spaces with underscores in the third column

        # Extract YouTube links from the subsequent columns
        links = []
        for cell in row[3:]:
            links.extend(extract_youtube_links_from_rows(cell))

        if links:
            for i, link in enumerate(links):
                if i == 0:
                    processed_data.append([col1, col2, col3, link])
                else:
                    processed_data.append([col1, col2, f"{col3}_{i + 1}", link])
        else:
            processed_data.append([col1, col2, col3, None])

    # Convert the processed data to a DataFrame
    processed_df = pd.DataFrame(processed_data,
                                columns=[sheet.cell(1, 1).value, sheet.cell(1, 2).value, sheet.cell(1, 3).value,
                                         'YouTube Links'])

    # Output the processed data to a new Excel sheet
    output_path = f'./output_files/{name}/processing/processed_video_links_v2.xlsx'
    with pd.ExcelWriter(output_path) as writer:
        processed_df.to_excel(writer, sheet_name='YouTube Links', index=False)

    print(f"Processed file saved to {output_path}")


def extract_youtube_links(file_path, name):
    if not os.path.exists(f'./output_files/{name}/processing/'):
        os.makedirs(f'./output_files/{name}/processing/')
    output_file=f'./output_files/{name}/processing/extracted_youtube_links.xlsx'
    youtube_links = []
    youtube_regex = r'(https?://(?:www\.)?youtube\.com/watch\?v=[\w-]+|https?://youtu\.be/[\w-]+)'

    # Load the file based on the extension
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
        for col in df.columns:
            for value in df[col].astype(str):
                matches = re.findall(youtube_regex, value)
                youtube_links.extend(matches)
    elif file_path.endswith('.xlsx'):
        wb = load_workbook(file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    # Check for hyperlinks
                    if cell.hyperlink and re.match(youtube_regex, cell.hyperlink.target):
                        youtube_links.append(cell.hyperlink.target)
                    # Check for plain text in cell value
                    elif isinstance(cell.value, str):
                        matches = re.findall(youtube_regex, cell.value)
                        youtube_links.extend(matches)
    else:
        raise ValueError("Unsupported file format. Please provide an Excel (.xlsx) or CSV (.csv) file.")

    # Save the links into a new CSV file
    links_df = pd.DataFrame({'Youtube Links': youtube_links})
    links_df.to_excel(output_file, index=False)

    print(f"Extracted {len(youtube_links)} YouTube links and saved to {output_file}")
    return output_file


def clean_title(title):
    return "".join([c for c in title if c.isalpha() or c.isdigit() or c==' ']).rstrip()


# Function to extract YouTube links from a hyperlink or plain text
def fetch_youtube_video_data(video_url):
    try:
        # Using yt-dlp to get video details
        with yt_dlp.YoutubeDL() as ydl:
            info = ydl.extract_info(video_url, download=False)

        video_data = {
            'title': info.get('title', ''),
            'upload_date': info.get('upload_date', ''),
            'views': info.get('view_count', 0),
            'duration': info.get('duration', 0),
            'uploader': info.get('uploader', ''),
            'like_count': info.get('like_count', 0),
            'comment_count': info.get('comment_count', 0)
        }
        print(f"Fetched data for {video_url}")
        return video_data

    except Exception as e:
        print(f"Error fetching data for {video_url}: {e}")
        return None


def update_excel_with_video_data(excel_file):
    # Load the Excel file into a DataFrame
    df = pd.read_excel(excel_file)

    # Check if the necessary columns exist; if not, create them
    columns_to_check = ['Upload Date', 'Views', 'Video Length (sec)', 'Uploaded By', 'Like Count', 'Comment Count']
    for column in columns_to_check:
        if column not in df.columns:
            df[column] = None

    # Iterate over the rows to fetch video data
    for index, row in df.iterrows():
        video_url = row['Youtube Links']  # Assuming 'Link' is the column name for URLs
        if pd.notna(video_url):  # Check if the URL is not NaN
            video_data = fetch_youtube_video_data(video_url)
            if video_data:
                df.at[index, 'Title'] = clean_title(video_data['title'])
                if(len(clean_title(video_data['title']))>30):
                    df.at[index, 'File_Name'] = clean_title(video_data['title'])[:30]+'.mp4'
                else:
                    df.at[index, 'File_Name'] = clean_title(video_data['title'])+'.mp4'
                df.at[index, 'Upload Date'] = video_data['upload_date']
                df.at[index, 'Views'] = video_data['views']
                df.at[index, 'Video Length (sec)'] = video_data['duration']
                df.at[index, 'Uploaded By'] = video_data['uploader']
                df.at[index, 'Like Count'] = video_data['like_count']
                df.at[index, 'Comment Count'] = video_data['comment_count']


    # Save the updated DataFrame back to the Excel file
    df.to_excel(excel_file, index=False)
    print(f"Updated Excel file saved: {excel_file}")
    return excel_file


def download_youtube_video(url, output_path,name):
    downloaded=False
    try:
        yt = YouTube(url)
        file_name = f"{name}.mp4"
        file_path = os.path.join(output_path, file_name)

        # Check if the file already exists
        if os.path.exists(file_path):
            print(f"{file_name} already exists in {output_path}. Skipping download.")
        else:
            print(f"Downloading {yt.title}")
            video_stream = yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution').first()
            video_stream.download(output_path=output_path, filename=file_name)
            print(f"Downloaded {file_name} to {output_path}")
            downloaded=True

    except Exception as e:
        print(f"An error occurred: {e}")
    return downloaded


def downloader(file_path,folder_name):
    if os.path.exists(f'./videos/{folder_name}/'):
        pass
    else:
        os.makedirs(f'./videos/{folder_name}/')
    df=pd.read_excel(file_path)
    for index,row in df.iterrows():
        link=row['Youtube Links']
        name=row['File_Name']
        if link:
            downloaded=download_youtube_video(link, output_path=f'./videos/{folder_name}/',name=name[:-4])
            row['Downloaded']=downloaded
        else:
            row['Downloaded']=False
    df.to_excel(file_path,index=False)


def merge_data(excel_file1, excel_file2,project_name):
    # Load the two Excel sheets
    df1 = pd.read_excel(excel_file1)
    df2 = pd.read_excel(excel_file2)

    # Merge the two DataFrames on the 'Title' column
    merged_data = pd.merge(df1, df2, how='inner', on='File_Name')

    # Save the merged DataFrame to a new Excel file
    merged_data.to_excel(f'./output_files/{project_name}/{project_name}_merged_videos_data.xlsx', index=False)

    print("Merge complete. Saved in output folder.")
if __name__ == '__main__':
    # Load the Excel file containing video links
    # original_file_path = './input_data/CMO/CMO_dataset.xlsx'
    folder_name='CMO'
    #processed_file=process_video_links(original_file_path,folder_name)
    # processed_file=update_excel_with_video_data(processed_file)
    # print("All videos data fetched.")
    # print(f"Processed file saved to {processed_file}")
    # processed_file='./output_files/processing/CMO/Processed_Video_Links.xlsx'
    # downloader(processed_file,folder_name)
    # print("All videos downloaded.")
    videos_folder_path = f'./videos/{folder_name}/'
    os.makedirs(f'./output_files/{folder_name}/', exist_ok=True)
    output_excel = f'./output_files/{folder_name}/{folder_name}_video_analysis_results.xlsx'
    analyze_folder(videos_folder_path, output_excel)


