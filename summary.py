from __future__ import unicode_literals
from googleapiclient.discovery import build
from pytubefix import YouTube
from pytube.exceptions import RegexMatchError, VideoUnavailable, ExtractError
import youtube_dl
import openpyxl
import os
import yt_dlp
import pandas as pd

def clean_title(title):
    return "".join([c for c in title if c.isalpha() or c.isdigit() or c==' ']).rstrip()


def get_youtube_video_links(api_key, keyword, max_results):
    youtube = build('youtube', 'v3', developerKey=api_key)

    # Load nextPageToken if it exists
    next_page_token = None
    if os.path.exists('nextpage.txt'):
        with open('nextpage.txt', 'r') as file:
            next_page_token = file.read().strip()

    video_links = []

    while True:
        search_request = youtube.search().list(
            part='snippet',
            q=keyword,
            type='video',
            maxResults=max_results,
            pageToken=next_page_token
        )

        search_response = search_request.execute()

        for item in search_response['items']:
            video_id = item['id']['videoId']
            video_title = item['snippet']['title']
            video_url = f"https://www.youtube.com/watch?v={video_id}"
            video_links.append({
                'Title': clean_title(video_title),
                'Link': video_url
            })

        # Check for the next page token and save it to a file
        next_page_token = search_response.get('nextPageToken', None)
        if next_page_token:
            with open('nextpage.txt', 'w') as file:
                file.write(next_page_token)
        else:
            # If there's no next page, remove the token file
            if os.path.exists('nextpage.txt'):
                os.remove('nextpage.txt')
            break

    return video_links


def save_video_links_to_excel(api_key,keyword,max_results,filename='youtube_video_links.xlsx'):

    video_links= get_youtube_video_links(api_key, keyword, max_results)
    # Load existing workbook or create a new one
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title='Video Links')

        # Create headers for the new sheet
        headers = ['Title', 'Link']
        sheet.append(headers)

    sheet = workbook['Video Links']

    # Add video links
    for video in video_links:
        sheet.append([video['Title'], video['Link']])

    # Save the workbook
    workbook.save(filename)


def download_youtube_video(url, output_path):
    try:
        yt = YouTube(url)
        video_title = yt.title
        sanitized_title = "".join([c for c in video_title if c.isalpha() or c.isdigit() or c==' ']).rstrip()
        file_name = f"{sanitized_title}.mp4"
        file_path = os.path.join(output_path, file_name)

        # Check if the file already exists
        if os.path.exists(file_path):
            print(f"{file_name} already exists in {output_path}. Skipping download.")
        else:
            print(f"Downloading {yt.title}")
            video_stream = yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution').first()
            video_stream.download(output_path)
            print(f"Downloaded {file_name} to {output_path}")

    except Exception as e:
        print(f"An error occurred: {e}")


def downloader(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, link = row
        download_youtube_video(link, output_path='C:\\Users\\Pavan\\PycharmProjects\\youtube_data\\videos\\')

    print("All videos downloaded.")


def fetch_youtube_video_data(video_url):
    try:
        # Using yt-dlp to get video details
        with yt_dlp.YoutubeDL() as ydl:
            info = ydl.extract_info(video_url, download=False)

        video_data = {
            'title': info.get('title', ''),
            'upload_date': info.get('upload_date', ''),
            'views': info.get('view_count', 0),
            'duration': info.get('duration', 0),
            'uploader': info.get('uploader', ''),
            'like_count': info.get('like_count', 0),
            'comment_count': info.get('comment_count', 0)
        }
        print(f"Fetched data for {video_url}")
        return video_data

    except Exception as e:
        print(f"Error fetching data for {video_url}: {e}")
        return None


def update_excel_with_video_data(excel_file):
    # Load the Excel file into a DataFrame
    df = pd.read_excel(excel_file)

    # Check if the necessary columns exist; if not, create them
    columns_to_check = ['Upload Date', 'Views', 'Video Length (sec)', 'Uploaded By', 'Like Count', 'Comment Count']
    for column in columns_to_check:
        if column not in df.columns:
            df[column] = None

    # Iterate over the rows to fetch video data
    for index, row in df.iterrows():
        video_url = row['Link']  # Assuming 'Link' is the column name for URLs
        if pd.notna(video_url):  # Check if the URL is not NaN
            video_data = fetch_youtube_video_data(video_url)
            if video_data:
                df.at[index, 'Upload Date'] = video_data['upload_date']
                df.at[index, 'Views'] = video_data['views']
                df.at[index, 'Video Length (sec)'] = video_data['duration']
                df.at[index, 'Uploaded By'] = video_data['uploader']
                df.at[index, 'Like Count'] = video_data['like_count']
                df.at[index, 'Comment Count'] = video_data['comment_count']

    # Save the updated DataFrame back to the Excel file
    df.to_excel(excel_file, index=False)
    print(f"Updated Excel file saved: {excel_file}")


if __name__ == '__main__':
    # Example usage
    # api_key = '***REMOVED***'
    # save_video_links_to_excel(video_links)
    # downloader('youtube_video_links.xlsx')
    update_excel_with_video_data('youtube_video_links.xlsx')


