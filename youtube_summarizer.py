import os
import re
import time
import pandas as pd
import yt_dlp
import openpyxl
import streamlit as st
from pytubefix import YouTube
from googleapiclient.discovery import build
import google.generativeai as genai
import json
import dotenv


# Load environment variables
dotenv.load_dotenv()

# Configure API keys securely
YOUTUBE_API_KEY = os.getenv('YOUTUBE_API_KEY')
GENAI_API_KEY = os.getenv('GENAI_API_KEY')

def clean_title(title):
    """Cleans the video title by removing unwanted characters."""
    return "".join([c for c in title if c.isalpha() or c.isdigit() or c == ' ']).rstrip()

def get_youtube_video_links(api_key, keyword, max_results):
    try:
        youtube = build('youtube', 'v3', developerKey=api_key)

        video_links = []
        next_page_token = None
        total_fetched = 0

        while total_fetched < max_results:
            search_request = youtube.search().list(
                part='snippet',
                q=keyword,
                type='video',
                maxResults=min(50, max_results - total_fetched),
                pageToken=next_page_token
            )
            search_response = search_request.execute()

            for item in search_response['items']:
                video_id = item['id']['videoId']
                video_title = item['snippet']['title']
                video_url = f"https://www.youtube.com/watch?v={video_id}"
                video_links.append({
                    'Title': clean_title(video_title),
                    'Link': video_url
                })
                total_fetched += 1
                if total_fetched >= max_results:
                    break

            next_page_token = search_response.get('nextPageToken')
            if not next_page_token:
                break

        return video_links

    except Exception as e:
        st.error(f"An error occurred while fetching video links: {e}")
        return []

def save_video_links_to_excel(api_key, keyword, max_results):
    try:
        filename = f'./output_files/{keyword}/processing/youtube_video_links.xlsx'
        os.makedirs(os.path.dirname(filename), exist_ok=True)

        video_links = get_youtube_video_links(api_key, keyword, max_results)
        if not video_links:
            st.error("No video links were retrieved.")
            return

        # Create a new workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Video Links'

        # Create headers for the new sheet
        headers = ['Title', 'Link']
        sheet.append(headers)

        # Add video links
        for video in video_links:
            sheet.append([video['Title'], video['Link']])

        # Save the workbook
        workbook.save(filename)
        st.success(f"Video links saved to {filename}")

    except Exception as e:
        st.error(f"An error occurred while saving video links: {e}")

def extract_youtube_links(file_path, name):
    try:
        output_file = f'./output_files/{name}/processing/extracted_youtube_links.xlsx'
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        youtube_links = []
        youtube_regex = r'(https?://(?:www\.)?youtube\.com/watch\?v=[\w-]+|https?://youtu\.be/[\w-]+)'

        # Load the file based on the extension
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
            for col in df.columns:
                for value in df[col].astype(str):
                    matches = re.findall(youtube_regex, value)
                    youtube_links.extend(matches)
        elif file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        # Check for hyperlinks
                        if cell.hyperlink and re.match(youtube_regex, cell.hyperlink.target):
                            youtube_links.append(cell.hyperlink.target)
                        # Check for plain text in cell value
                        elif isinstance(cell.value, str):
                            matches = re.findall(youtube_regex, cell.value)
                            youtube_links.extend(matches)
        else:
            st.error("Unsupported file format. Please provide an Excel (.xlsx) or CSV (.csv) file.")
            return None

        # Save the links into a new Excel file
        links_df = pd.DataFrame({'Youtube Links': youtube_links})
        links_df.to_excel(output_file, index=False)
        st.success(f"Extracted {len(youtube_links)} YouTube links and saved to {output_file}")
        return output_file

    except Exception as e:
        st.error(f"An error occurred while extracting YouTube links: {e}")
        return None

def fetch_youtube_video_data(video_url):
    try:
        # Using yt-dlp to get video details
        with yt_dlp.YoutubeDL() as ydl:
            info = ydl.extract_info(video_url, download=False)

        video_data = {
            'title': info.get('title', ''),
            'upload_date': info.get('upload_date', ''),
            'views': info.get('view_count', 0),
            'duration': info.get('duration', 0),
            'uploader': info.get('uploader', ''),
            'like_count': info.get('like_count', 0),
            'comment_count': info.get('comment_count', 0)
        }
        return video_data

    except Exception as e:
        st.error(f"Error fetching data for {video_url}: {e}")
        return None

def update_excel_with_video_data(excel_file):
    try:
        # Load the Excel file into a DataFrame
        df = pd.read_excel(excel_file)

        # Check if the necessary columns exist; if not, create them
        columns_to_check = ['Upload Date', 'Views', 'Video Length (sec)', 'Uploaded By', 'Like Count', 'Comment Count', 'File_Name']
        for column in columns_to_check:
            if column not in df.columns:
                df[column] = None

        # Iterate over the rows to fetch video data
        for index, row in df.iterrows():
            video_url = row['Youtube Links']  # Assuming 'Youtube Links' is the column name for URLs
            if pd.notna(video_url):  # Check if the URL is not NaN
                video_data = fetch_youtube_video_data(video_url)
                if video_data:
                    df.at[index, 'Title'] = clean_title(video_data['title'])
                    file_name = f"{clean_title(video_data['title'])[:30]}.mp4"
                    df.at[index, 'File_Name'] = file_name
                    df.at[index, 'Upload Date'] = video_data['upload_date']
                    df.at[index, 'Views'] = video_data['views']
                    df.at[index, 'Video Length (sec)'] = video_data['duration']
                    df.at[index, 'Uploaded By'] = video_data['uploader']
                    df.at[index, 'Like Count'] = video_data['like_count']
                    df.at[index, 'Comment Count'] = video_data['comment_count']

        # Save the updated DataFrame back to the Excel file
        df.to_excel(excel_file, index=False)
        st.success(f"Updated Excel file saved: {excel_file}")
        return excel_file

    except Exception as e:
        st.error(f"An error occurred while updating Excel with video data: {e}")
        return None

def download_youtube_video(url, output_path, name):
    try:
        yt = YouTube(url, client="ANDROID_VR") # Use the 'WEB_CREATOR' user agent to bypass age restrictions
        file_name = f"{name}.mp4"
        file_path = os.path.join(output_path, file_name)

        # Check if the file already exists
        if os.path.exists(file_path):
            st.write(f"{file_name} already exists in {output_path}. Skipping download.")
            return True
        else:
            st.write(f"Downloading {yt.title}")
            video_stream = yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution').first()
            video_stream.download(output_path=output_path, filename=file_name)
            st.write(f"Downloaded {file_name} to {output_path}")
            return True

    except Exception as e:
        print(e)
        st.error(f"An error occurred while downloading video: {e}")
        return False

def downloader(file_path, folder_name):
    try:
        download_path = f'./videos/{folder_name}/'
        os.makedirs(download_path, exist_ok=True)
        df = pd.read_excel(file_path)

        df['Downloaded'] = False  # Initialize 'Downloaded' column

        for index, row in df.iterrows():
            link = row['Youtube Links']
            name = row['File_Name'].replace('.mp4', '')
            if link:
                downloaded = download_youtube_video(link, output_path=download_path, name=name)
                df.at[index, 'Downloaded'] = downloaded
            else:
                df.at[index, 'Downloaded'] = False

        df.to_excel(file_path, index=False)
        st.success("All videos have been processed for download.")
    except Exception as e:
        st.error(f"An error occurred during video downloading: {e}")

def merge_data(excel_file1, excel_file2, project_name):
    try:
        # Load the two Excel sheets
        df1 = pd.read_excel(excel_file1)
        df2 = pd.read_excel(excel_file2)

        # Merge the two DataFrames on the 'File_Name' column
        merged_data = pd.merge(df1, df2, how='inner', on='File_Name')

        # Save the merged DataFrame to a new Excel file
        output_file = f'./output_files/{project_name}/{project_name}_merged_videos_data.xlsx'
        merged_data.to_excel(output_file, index=False)
        st.success(f"Merge complete. Saved as {output_file}")
    except Exception as e:
        st.error(f"An error occurred while merging data: {e}")

def upload_to_gemini(path, mime_type=None):
    try:
        file = genai.upload_file(path, mime_type=mime_type)
        return file
    except Exception as e:
        st.error(f"An error occurred while uploading to Gemini: {e}")
        return None

def wait_for_file_active(file):
    try:
        st.write("Waiting for file processing...")
        file_status = genai.get_file(file.name)
        while file_status.state.name == "PROCESSING":
            st.write(".", end="", flush=True)
            time.sleep(10)
            file_status = genai.get_file(file.name)
        if file_status.state.name != "ACTIVE":
            raise Exception(f"File {file.name} failed to process")
        st.write("File is ready.")
        return file_status
    except Exception as e:
        st.error(f"An error occurred while waiting for file to become active: {e}")
        return None

def analyze_video(video_path):
    try:
        st.write(f"Uploading video file: {video_path}")
        file = upload_to_gemini(video_path, mime_type="video/mp4")
        if not file:
            return None

        st.write("Waiting for file to be ready...")
        file_status = wait_for_file_active(file)
        if not file_status:
            return None

        st.write("Starting chat session...")
        chat_session = genai.GenerativeModel(
            model_name="gemini-1.5-flash",
            generation_config={
                "temperature": 1,
                "top_p": 0.95,
                "top_k": 64,
                "max_output_tokens": 8192,
                "response_mime_type": "application/json",
            },
        ).start_chat(history=[{
          "role": "user",
          "parts": [
            file,
          ],
        },])

        prompt = '''For the given video, analyze the video and provide me the following details in JSON format:
        "details": {
                "video_summary": "Provide a clear descriptive summary of the video's content.",
                "number_of_characters": "Count the number of distinct characters or people appearing in the video that have impact on the video.",
                "video_total_emotion": "Analyze the overall emotional tone of the video (e.g., positive, negative, neutral).",
                "video_duration": "Provide the duration of the video in seconds.",
                "video_language": "Identify the primary language spoken in the video.",
                "video_genre": "Identify the genre or category of the video (e.g., comedy, drama, action).",
                "video_mood": "Identify the mood or atmosphere of the video (e.g., suspenseful, romantic, humorous).",
                "video_tone": "Identify the tone or style of the video (e.g., serious, light-hearted, satirical).",
                "video_setting": "Identify the primary setting or location of the video.",
                "video_theme": "Identify the central theme or message of the video.",
                "video_impact": "Analyze the impact or significance of the video on the viewer.",
                "video_target_audience": "Identify the target audience or demographic for the video.",
                "video_influences": "Identify any notable influences or inspirations for the video.",
                "video_references": "Identify any references or allusions made in the video.",
                "color_palette": "Identify the primary color palette used in the video.",
                "words_used": "Identify any key words or phrases that are repeated or emphasized in the video.",
                "brand_name": "Identify the brand name associated with the video, especially if it's an advertisement.",
                "product_name": "Identify the product or service being promoted in the video.",
                "product_features": "Identify the key features or benefits of the product or service.",
                "product_target_audience": "Identify the target audience or demographic for the product or service.",
                "product_message": "Identify the central message or value proposition of the product or service.",
                "product_call_to_action": "Identify the call-to-action or desired response from the viewer.",
                "Personality Trait": "Identify the personality of the person in the video. Identify in these categories:  Extraversion, Agreeableness, Neuroticism, Conscientiousness, Openness to Experience, Narcissism, Machiavellianism, Psychopathy, Maverickism",
                "Terms Associated": "Identify the terms associated with personality traits",
                "This person will be": "Identify the personality of the person in the video Example-(Will demonstrate energy, action, assertiveness in his words and actions), (Will act whimsically, without any consideration of surrounding or others. They have a superficial charm that lures people, but they lack emotions or empathy)",
                "Person Tone": "Identify the tone of the person in the video. Example-(Serious, Light-hearted, Satirical)",
                "Person Mood": "Identify the mood of the person in the video. Example-(Suspenseful, Romantic, Humorous)",
                "Person emotion": "Identify the emotion of the person in the video. Example-(Positive, Negative, Neutral)"
            }'''

        response = chat_session.send_message(prompt)
        data = json.loads(response.text)
        return data
    except Exception as e:
        st.error(f"An error occurred during video analysis: {e}")
        return None

def analyze_folder(folder_path, project_name):
    try:
        output_excel = f'./output_files/{project_name}/processing/ai_analyzed_video_data.xlsx'
        os.makedirs(os.path.dirname(output_excel), exist_ok=True)
        video_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.mp4')]
        all_data = []

        for video in video_files:
            st.write(f"Processing video: {video}")
            video_data = analyze_video(video)
            if video_data:
                video_data['File_Name'] = os.path.basename(video)
                all_data.append(video_data)
                st.write(f"Analysis complete for {video}")
            else:
                st.error(f"Analysis failed for {video}")
            time.sleep(60)  # To respect API rate limits

        # Convert the collected data into a DataFrame and save it as an Excel file
        df = pd.DataFrame(all_data)
        df.to_excel(output_excel, index=False)
        st.success(f"Analysis complete. Results saved to {output_excel}")
        return output_excel
    except Exception as e:
        st.error(f"An error occurred during folder analysis: {e}")
        return None

def app():
    # Securely get the API keys from environment variables
    YOUTUBE_API_KEY = os.getenv('YOUTUBE_API_KEY')
    GENAI_API_KEY = os.getenv('GENAI_API_KEY')

    if not YOUTUBE_API_KEY:
        st.error("YouTube API key not found. Please set the 'YOUTUBE_API_KEY' environment variable.")
        return
    if not GENAI_API_KEY:
        st.error("Generative AI API key not found. Please set the 'GENAI_API_KEY' environment variable.")
        return

    # Configure the Generative AI API
    genai.configure(api_key=GENAI_API_KEY)

    option = st.sidebar.radio('Select an option', ['Keyword Search', 'Single Video', 'File Mode'])

    if option == 'Keyword Search':
        st.title('YouTube Video Summarizer - Keyword Search')
        keyword = st.text_input('Enter the keyword to search')
        max_results = st.slider('Select the number of videos to search', min_value=1, max_value=50)
        if st.button('Search'):
            if keyword:
                save_video_links_to_excel(YOUTUBE_API_KEY, keyword, int(max_results))
                project_name = keyword.replace(' ', '_')
                file = f'./output_files/{keyword}/processing/youtube_video_links.xlsx'
                if file and project_name:
                    try:
                        st.write('Extracting YouTube links from the file...')
                        excel_file = extract_youtube_links(file, project_name)
                        if excel_file:
                            st.write('Fetching video data from Excel file links...')
                            updated_excel = update_excel_with_video_data(excel_file)
                            if updated_excel:
                                st.write('Downloading videos...')
                                downloader(updated_excel, project_name)
                                st.write('All videos downloaded.')
                                videos_folder_path = f'./videos/{project_name}/'
                                st.write('Analyzing videos...')
                                summary_file = analyze_folder(videos_folder_path, project_name)
                                if summary_file:
                                    st.write(f'All videos analyzed, saved to {summary_file}')
                                    st.write('Merging data...')
                                    merge_data(updated_excel, summary_file, project_name)
                                    st.write('All data merged and saved to output folder.')
                                else:
                                    st.error("Video analysis failed.")
                            else:
                                st.error("Failed to update Excel with video data.")
                        else:
                            st.error("Failed to extract YouTube links.")
                    except Exception as e:
                        st.error(f"An error occurred during file processing: {e}")
                else:
                    st.error("Please enter valid file path and project name.")

            else:
                st.error("Please enter a keyword to search.")

    elif option == 'Single Video':
        st.title('YouTube Video Summarizer - Single Video')
        url = st.text_input('Enter the video link')
        video_path = './videos/single_url/'
        if st.button('Summarize'):
            if url:
                try:
                    st.write('Fetching video data...')
                    data = fetch_youtube_video_data(url)
                    if data:
                        file_name = clean_title(data['title'])[:30]
                        st.write('Downloading video...')
                        downloaded = download_youtube_video(url, video_path, file_name)
                        video_full_path = os.path.join(video_path, f"{file_name}.mp4")
                        if os.path.exists(video_full_path):
                            st.write(f'Video saved to {video_full_path}')
                            st.write('Analyzing the video...')
                            output = analyze_video(video_full_path)
                            if output:
                                st.json(output)
                            else:
                                st.error("Video analysis failed.")
                        else:
                            st.error("Video download failed.")
                    else:
                        st.error("Failed to fetch video data.")
                except Exception as e:
                    st.error(f"An error occurred: {e}")
            else:
                st.error("Please enter a video URL.")

    elif option == 'File Mode':
        st.title('YouTube Video Summarizer - File Mode')
        file = st.text_input('Enter the file path')
        project_name = st.text_input('Enter the project folder name')
        if st.button('Extract Links'):
            if file and project_name:
                try:
                    st.write('Extracting YouTube links from the file...')
                    excel_file = extract_youtube_links(file, project_name)
                    if excel_file:
                        st.write('Fetching video data from Excel file links...')
                        updated_excel = update_excel_with_video_data(excel_file)
                        if updated_excel:
                            st.write('Downloading videos...')
                            downloader(updated_excel, project_name)
                            st.write('All videos downloaded.')
                            videos_folder_path = f'./videos/{project_name}/'
                            st.write('Analyzing videos...')
                            summary_file = analyze_folder(videos_folder_path, project_name)
                            if summary_file:
                                st.write(f'All videos analyzed, saved to {summary_file}')
                                st.write('Merging data...')
                                merge_data(updated_excel, summary_file, project_name)
                                st.write('All data merged and saved to output folder.')
                            else:
                                st.error("Video analysis failed.")
                        else:
                            st.error("Failed to update Excel with video data.")
                    else:
                        st.error("Failed to extract YouTube links.")
                except Exception as e:
                    st.error(f"An error occurred during file processing: {e}")
            else:
                st.error("Please enter valid file path and project name.")

if __name__ == '__main__':
    app()
