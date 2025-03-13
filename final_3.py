import os
import re
import time
import json
import dotenv
import logging
import pandas as pd
import openpyxl
import streamlit as st
from pytubefix import YouTube
from googleapiclient.discovery import build
import google.generativeai as genai
import yt_dlp

# Configure logging to terminal
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# Load environment variables
dotenv.load_dotenv()

# Configure API keys securely
YOUTUBE_API_KEY = os.getenv('YOUTUBE_API_KEY')
GENAI_API_KEY = os.getenv('GENAI_API_KEY')

def clean_title(title):
    """Cleans the video title by removing unwanted characters."""
    return "".join([c for c in title if c.isalpha() or c.isdigit() or c == ' ']).rstrip()

def get_youtube_video_links(api_key, keyword, max_results):
    video_links = []
    try:
        youtube = build('youtube', 'v3', developerKey=api_key)

        next_page_token = None
        total_fetched = 0

        while total_fetched < max_results:
            search_request = youtube.search().list(
                part='snippet',
                q=keyword,
                type='video',
                maxResults=min(50, max_results - total_fetched),
                pageToken=next_page_token
            )
            search_response = search_request.execute()

            for item in search_response.get('items', []):
                video_id = item['id']['videoId']
                video_title = item['snippet']['title']
                video_url = f"https://www.youtube.com/watch?v={video_id}"
                video_links.append({
                    'Title': clean_title(video_title),
                    'Link': video_url
                })
                total_fetched += 1
                if total_fetched >= max_results:
                    break

            next_page_token = search_response.get('nextPageToken')
            if not next_page_token:
                break
    except Exception as e:
        logger.exception("Error fetching video links.")
    return video_links

def save_video_links_to_excel(api_key, keyword, max_results):
    filename = f'./output_files/{keyword.replace(' ', '_')}/processing/youtube_video_links.xlsx'
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    video_links = get_youtube_video_links(api_key, keyword, max_results)
    if not video_links:
        st.warning("No video links were retrieved.")
        return None

    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Video Links'

    # Create headers for the new sheet
    headers = ['Title', 'Link']
    sheet.append(headers)

    # Add video links
    for video in video_links:
        sheet.append([video['Title'], video['Link']])

    # Save the workbook
    workbook.save(filename)
    return filename

def extract_youtube_links(file_path, name):
    output_file = f'./output_files/{name.replace(' ', '_')}/processing/extracted_youtube_links.xlsx'
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    youtube_links = []
    youtube_regex = r'(https?://(?:www\.)?youtube\.com/watch\?v=[\w-]+|https?://youtu\.be/[\w-]+)'

    try:
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
            for col in df.columns:
                for value in df[col].astype(str):
                    matches = re.findall(youtube_regex, value)
                    youtube_links.extend(matches)
        elif file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        # Check for hyperlinks
                        if cell.hyperlink and re.match(youtube_regex, cell.hyperlink.target):
                            youtube_links.append(cell.hyperlink.target)
                        elif isinstance(cell.value, str):
                            matches = re.findall(youtube_regex, cell.value)
                            youtube_links.extend(matches)
        else:
            logger.error("Unsupported file format.")
            return None

        links_df = pd.DataFrame({'Youtube Links': youtube_links})
        links_df.to_excel(output_file, index=False)
        return output_file
    except Exception as e:
        logger.exception("Error extracting YouTube links.")
        return None

def fetch_youtube_video_data(video_url):
    try:
        with yt_dlp.YoutubeDL() as ydl:
            info = ydl.extract_info(video_url, download=False)
        video_data = {
            'title': info.get('title', ''),
            'upload_date': info.get('upload_date', ''),
            'views': info.get('view_count', 0),
            'duration': info.get('duration', 0),
            'uploader': info.get('uploader', ''),
            'like_count': info.get('like_count', 0),
            'comment_count': info.get('comment_count', 0)
        }
        return video_data
    except Exception:
        logger.exception(f"Error fetching data for {video_url}")
        return None

def update_excel_with_video_data(excel_file):
    try:
        df = pd.read_excel(excel_file)

        columns_to_check = ['Title', 'Upload Date', 'Views', 'Video Length (sec)', 'Uploaded By', 'Like Count', 'Comment Count', 'File_Name']
        for column in columns_to_check:
            if column not in df.columns:
                df[column] = None

        for index, row in df.iterrows():
            video_url = row['Youtube Links'] if 'Youtube Links' in df.columns else None
            if pd.notna(video_url):
                video_data = fetch_youtube_video_data(video_url)
                if video_data:
                    cleaned_title = clean_title(video_data['title'])
                    file_name = f"{cleaned_title[:30]}.mp4"
                    df.at[index, 'Title'] = cleaned_title
                    df.at[index, 'File_Name'] = file_name
                    df.at[index, 'Upload Date'] = video_data['upload_date']
                    df.at[index, 'Views'] = video_data['views']
                    df.at[index, 'Video Length (sec)'] = video_data['duration']
                    df.at[index, 'Uploaded By'] = video_data['uploader']
                    df.at[index, 'Like Count'] = video_data['like_count']
                    df.at[index, 'Comment Count'] = video_data['comment_count']

        df.to_excel(excel_file, index=False)
        return excel_file
    except Exception:
        logger.exception("Error updating Excel with video data.")
        return None

def download_youtube_video(url, output_path, name):
    try:
        yt = YouTube(url, client="ANDROID_VR")
        file_name = f"{name}.mp4"
        file_path = os.path.join(output_path, file_name)

        if os.path.exists(file_path):
            logger.info(f"{file_name} already exists. Skipping download.")
            return True
        else:
            video_stream = yt.streams.filter(progressive=True, file_extension='mp4').order_by('resolution').first()
            if not video_stream:
                logger.error("No suitable video stream found.")
                return False
            video_stream.download(output_path=output_path, filename=file_name)
            return True
    except Exception:
        logger.exception(f"Error downloading video {url}")
        return False

def downloader(file_path, folder_name):
    download_path = f'./videos/{folder_name}/'
    os.makedirs(download_path, exist_ok=True)
    df = pd.read_excel(file_path)
    if 'Downloaded' not in df.columns:
        df['Downloaded'] = False

    total_videos = len(df)
    success_count = 0

    progress_bar = st.progress(0)
    status_text = st.empty()

    for index, row in df.iterrows():
        link = row['Youtube Links'] if 'Youtube Links' in df.columns else None
        name = (row['File_Name'].replace('.mp4', '') if 'File_Name' in df.columns and pd.notna(row['File_Name']) else f"video_{index}")
        if link:
            status_text.text(f"Downloading video {index+1}/{total_videos}")
            downloaded = download_youtube_video(link, output_path=download_path, name=name)
            df.at[index, 'Downloaded'] = downloaded
            if downloaded:
                success_count += 1
        else:
            df.at[index, 'Downloaded'] = False

        progress_bar.progress((index+1)/total_videos)

    df.to_excel(file_path, index=False)
    st.write(f"Downloaded {success_count}/{total_videos} videos.")
    return download_path

def merge_data(excel_file1, excel_file2, project_name):
    try:
        df1 = pd.read_excel(excel_file1)
        df2 = pd.read_excel(excel_file2)

        merged_data = pd.merge(df1, df2, how='inner', on='File_Name')
        output_file = f'./output_files/{project_name.replace(' ', '_')}/{project_name}_merged_videos_data.xlsx'
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        merged_data.to_excel(output_file, index=False)
        return output_file
    except Exception:
        logger.exception("Error merging data.")
        return None

def upload_to_gemini(path, mime_type=None):
    try:
        file = genai.upload_file(path, mime_type=mime_type)
        return file
    except Exception:
        logger.exception(f"Error uploading {path} to Gemini.")
        return None

def wait_for_file_active(file):
    try:
        file_status = genai.get_file(file.name)
        while file_status.state.name == "PROCESSING":
            time.sleep(10)
            file_status = genai.get_file(file.name)
        if file_status.state.name != "ACTIVE":
            logger.error(f"File {file.name} failed to process.")
            return None
        return file_status
    except Exception:
        logger.exception("Error waiting for file to become active.")
        return None

def flatten_json(json_data):
    flat_data = {}
    for key, value in json_data.items():
        if isinstance(value, dict):
            for sub_key, sub_value in flatten_json(value).items():
                flat_data[f"{key}_{sub_key}"] = sub_value
        else:
            flat_data[key] = value
    return flat_data

def analyze_video(video_path):
    try:
        file = upload_to_gemini(video_path, mime_type="video/mp4")
        if not file:
            return None

        file_status = wait_for_file_active(file)
        if not file_status:
            return None

        chat_session = genai.GenerativeModel(
            model_name="gemini-1.5-flash",
            generation_config={
                "temperature": 1,
                "top_p": 0.95,
                "top_k": 64,
                "max_output_tokens": 8192,
                "response_mime_type": "application/json",
            },
        ).start_chat(history=[{
            "role": "user",
            "parts": [
                file,
            ],
        },])

        prompt = '''For the given video, analyze the video and provide me the following details in JSON format:
                "details": {
                        "video_summary": "Provide a clear descriptive summary of the video's content.",
                        "number_of_characters": "Count the number of distinct characters or people appearing in the video that have impact on the video.",
                        "video_total_emotion": "Analyze the overall emotional tone of the video (e.g., positive, negative, neutral).",
                        "video_duration": "Provide the duration of the video in seconds.",
                        "video_language": "Identify the primary language spoken in the video.",
                        "video_genre": "Identify the genre or category of the video (e.g., comedy, drama, action).",
                        "video_mood": "Identify the mood or atmosphere of the video (e.g., suspenseful, romantic, humorous).",
                        "video_tone": "Identify the tone or style of the video (e.g., serious, light-hearted, satirical).",
                        "video_setting": "Identify the primary setting or location of the video.",
                        "video_theme": "Identify the central theme or message of the video.",
                        "video_impact": "Analyze the impact or significance of the video on the viewer.",
                        "video_target_audience": "Identify the target audience or demographic for the video.",
                        "video_influences": "Identify any notable influences or inspirations for the video.",
                        "video_references": "Identify any references or allusions made in the video.",
                        "color_palette": "Identify the primary color palette used in the video.",
                        "words_used": "Identify any key words or phrases that are repeated or emphasized in the video.",
                        "brand_name": "Identify the brand name associated with the video, especially if it's an advertisement.",
                        "product_name": "Identify the product or service being promoted in the video.",
                        "product_features": "Identify the key features or benefits of the product or service.",
                        "product_target_audience": "Identify the target audience or demographic for the product or service.",
                        "product_message": "Identify the central message or value proposition of the product or service.",
                        "product_call_to_action": "Identify the call-to-action or desired response from the viewer.",
                        "Personality Trait": "Identify the personality of the person in the video. Identify in these categories:  Extraversion, Agreeableness, Neuroticism, Conscientiousness, Openness to Experience, Narcissism, Machiavellianism, Psychopathy, Maverickism",
                        "Terms Associated": "Identify the terms associated with personality traits",
                        "This person will be": "Identify the personality of the person in the video Example-(Will demonstrate energy, action, assertiveness in his words and actions), (Will act whimsically, without any consideration of surrounding or others. They have a superficial charm that lures people, but they lack emotions or empathy)",
                        "Person Tone": "Identify the tone of the person in the video. Example-(Serious, Light-hearted, Satirical)",
                        "Person Mood": "Identify the mood of the person in the video. Example-(Suspenseful, Romantic, Humorous)",
                        "Person emotion": "Identify the emotion of the person in the video. Example-(Positive, Negative, Neutral)"
                    }'''

        response = chat_session.send_message(prompt)
        data = json.loads(response.text)
        return data
    except Exception:
        logger.exception(f"Error analyzing video: {video_path}")
        return None

def analyze_folder(folder_path, project_name):
    output_excel = f'./output_files/{project_name.replace(' ', '_')}/processing/ai_analyzed_video_data.xlsx'
    os.makedirs(os.path.dirname(output_excel), exist_ok=True)
    video_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.mp4')]
    all_data = []
    total_videos = len(video_files)
    success_count = 0

    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, video in enumerate(video_files):
        status_text.text(f"Analyzing video {i+1}/{total_videos}")
        video_data = analyze_video(video)
        if video_data:
            flat_video_data = flatten_json(video_data)
            flat_video_data['File_Name'] = os.path.basename(video)
            all_data.append(flat_video_data)
            success_count += 1
        else:
            logger.error(f"Analysis failed for {video}. Continuing with next video.")
        progress_bar.progress((i+1)/total_videos)
        time.sleep(10)  # To respect any API rate limits

    df = pd.DataFrame(all_data)
    df.to_excel(output_excel, index=False)
    st.write(f"Analyzed {success_count}/{total_videos} videos successfully.")
    return output_excel

def app():
    # Set up a single status text element at the start


    if not YOUTUBE_API_KEY:
        st.warning("YouTube API key not found.")
        return
    if not GENAI_API_KEY:
        st.warning("Generative AI API key not found.")
        return

    genai.configure(api_key=GENAI_API_KEY)

    option = st.sidebar.radio('Select an option', ['Keyword Search', 'Single Video', 'File Mode'])

    if option == 'Keyword Search':
        st.title('YouTube Video Summarizer - Keyword Search')
        keyword = st.text_input('Enter the keyword to search')
        max_results = st.slider('Number of videos to fetch', min_value=1, max_value=50)
        if st.button('Search'):
            status_text = st.empty()
            if keyword:
                status_text.text("Fetching video links...")
                file = save_video_links_to_excel(YOUTUBE_API_KEY, keyword, int(max_results))
                if file:
                    project_name = keyword.replace(' ', '_')
                    status_text.text("Extracting YouTube links...")
                    excel_file = extract_youtube_links(file, project_name)
                    if excel_file:
                        status_text.text("Updating video data...")
                        updated_excel = update_excel_with_video_data(excel_file)
                        if updated_excel:
                            status_text.text("Downloading videos...")
                            videos_folder_path = downloader(updated_excel, project_name)
                            status_text.text("Analyzing videos...")
                            summary_file = analyze_folder(videos_folder_path, project_name)
                            if summary_file:
                                status_text.text("Merging data...")
                                merged_file = merge_data(updated_excel, summary_file, project_name)
                                if merged_file:
                                    status_text.text("All steps completed successfully.")
                                    st.success("Process complete!")
                                else:
                                    status_text.text("Merging encountered issues.")
                                    st.warning("Check logs for details.")
                            else:
                                status_text.text("Video analysis encountered issues.")
                                st.warning("Check logs for details.")
                        else:
                            status_text.text("Could not update Excel with video data.")
                            st.warning("Check logs for details.")
                    else:
                        status_text.text("Could not extract YouTube links.")
                        st.warning("Check logs for details.")
                else:
                    status_text.text("No videos fetched.")
                    st.warning("Check logs for details.")
            else:
                st.warning("Please enter a keyword.")

    elif option == 'Single Video':
        st.title('YouTube Video Summarizer - Single Video')
        url = st.text_input('Enter the video link')
        if st.button('Summarize'):
            status_text = st.empty()
            if url:
                status_text.text("Fetching video data...")
                data = fetch_youtube_video_data(url)
                if data:
                    file_name = clean_title(data['title'])[:30]
                    video_path = './videos/single_url/'
                    os.makedirs(video_path, exist_ok=True)
                    status_text.text("Downloading video...")
                    downloaded = download_youtube_video(url, video_path, file_name)
                    if downloaded:
                        status_text.text("Analyzing video...")
                        video_full_path = os.path.join(video_path, f"{file_name}.mp4")
                        if os.path.exists(video_full_path):
                            output = analyze_video(video_full_path)
                            if output:
                                status_text.text("Analysis complete.")
                                st.success("Analysis Result:")
                                st.json(output)
                            else:
                                status_text.text("Analysis failed.")
                                st.warning("Check logs for details.")
                        else:
                            status_text.text("Downloaded video not found.")
                            st.warning("Check logs for details.")
                    else:
                        status_text.text("Video download encountered issues.")
                        st.warning("Check logs for details.")
                else:
                    status_text.text("Could not fetch video data.")
                    st.warning("Check logs for details.")
            else:
                st.warning("Please enter a video URL.")

    elif option == 'File Mode':
        st.title('YouTube Video Summarizer - File Mode')
        file = st.text_input('Enter the file path (Excel or CSV)')
        project_name = st.text_input('Enter the project folder name')
        if st.button('Extract Links'):
            status_text = st.empty()
            if file and project_name:
                status_text.text("Extracting YouTube links...")
                excel_file = extract_youtube_links(file, project_name)
                if excel_file:
                    status_text.text("Updating video data...")
                    updated_excel = update_excel_with_video_data(excel_file)
                    if updated_excel:
                        status_text.text("Downloading videos...")
                        videos_folder_path = downloader(updated_excel, project_name)
                        status_text.text("Analyzing videos...")
                        summary_file = analyze_folder(videos_folder_path, project_name)
                        if summary_file:
                            status_text.text("Merging data...")
                            merged_file = merge_data(updated_excel, summary_file, project_name)
                            if merged_file:
                                status_text.text("All steps completed successfully.")
                                st.success("Process complete!")
                            else:
                                status_text.text("Merging encountered issues.")
                                st.warning("Check logs for details.")
                        else:
                            status_text.text("Video analysis encountered issues.")
                            st.warning("Check logs for details.")
                    else:
                        status_text.text("Could not update Excel with video data.")
                        st.warning("Check logs for details.")
                else:
                    status_text.text("Could not extract links.")
                    st.warning("Check logs for details.")
            else:
                st.warning("Please enter valid file path and project name.")

if __name__ == '__main__':
    app()