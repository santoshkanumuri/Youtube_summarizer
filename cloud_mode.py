import os
import re
import time
import json
import uuid
import dotenv
import logging
import pandas as pd
import openpyxl
import streamlit as st
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from googleapiclient.discovery import build
import google.generativeai as genai
import yt_dlp
import boto3
from botocore.exceptions import ClientError
from typing import Dict, List, Optional, Tuple

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# Load environment variables
dotenv.load_dotenv()

# Constants
MAX_RESULTS_PER_PAGE = 50
SLEEP_TIME_BETWEEN_ANALYSES = 10  # seconds
RETRY_ATTEMPTS = 3
BACKOFF_FACTOR = 2
TASKS_FILE = "./background_tasks.json"

# API & S3 Configurations
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
GENAI_API_KEY = os.getenv("GENAI_API_KEY")
S3_BUCKET = os.getenv("AWS_BUCKET_NAME")
S3_FOLDER = os.getenv("S3_FOLDER", "youtube_videos")
S3_OUTPUT_BUCKET = os.getenv("S3_OUTPUT_BUCKET", "youtube-summarizer-output-files")

# Global executor for background tasks
executor = ThreadPoolExecutor(max_workers=2)

### Utility Functions

def load_background_tasks() -> Dict[str, dict]:
    """Load background tasks from TASKS_FILE."""
    if os.path.exists(TASKS_FILE):
        try:
            with open(TASKS_FILE, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            logger.error(f"Failed to load tasks file: {e}")
            return {}
    return {}

def save_background_tasks(tasks: Dict[str, dict]) -> None:
    """Save background tasks to TASKS_FILE."""
    try:
        with open(TASKS_FILE, "w") as f:
            json.dump(tasks, f, indent=2)
    except IOError as e:
        logger.error(f"Failed to save tasks file: {e}")

def run_background_task(task_fn, *args, task_keyword: Optional[str] = None, **kwargs) -> str:
    """Run a function in the background and track its status."""
    task_id = str(uuid.uuid4())
    future = executor.submit(task_fn, *args, **kwargs)

    tasks = load_background_tasks()
    tasks[task_id] = {
        "task_id": task_id,
        "keyword": task_keyword,
        "status": "running",
        "result": None,
        "error": None,
        "timestamp": time.time()
    }
    save_background_tasks(tasks)

    def callback(future):
        tasks = load_background_tasks()
        try:
            result = future.result()
            tasks[task_id].update({"status": "completed", "result": result})
        except Exception as e:
            tasks[task_id].update({"status": "failed", "error": str(e)})
        save_background_tasks(tasks)

    future.add_done_callback(callback)
    return task_id

### S3 Interactions

def upload_videos_to_s3(directory: str, folder_name: Optional[str] = None) -> bool:
    """Upload .mp4 files to S3 and remove them locally after success."""
    s3 = boto3.client("s3")
    success = True
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(".mp4"):
                local_path = os.path.join(root, file)
                s3_key = f"{S3_FOLDER}/{folder_name}/{file}" if folder_name else f"{S3_FOLDER}/{file}"
                try:
                    s3.upload_file(local_path, S3_BUCKET, s3_key)
                    logger.info(f"Uploaded {file} to s3://{S3_BUCKET}/{s3_key}")
                    os.remove(local_path)
                except ClientError as e:
                    logger.error(f"Failed to upload {file} to S3: {e}")
                    success = False
    return success

def upload_final_outputs_to_s3(file_path: str) -> bool:
    """Upload final output file to S3 output bucket."""
    s3 = boto3.client("s3")
    file_name = os.path.basename(file_path).replace(" ", "_")
    try:
        s3.upload_file(file_path, S3_OUTPUT_BUCKET, file_name)
        logger.info(f"Uploaded {file_name} to s3://{S3_OUTPUT_BUCKET}/{file_name}")
        return True
    except ClientError as e:
        logger.error(f"Failed to upload {file_name} to S3: {e}")
        return False

### YouTube Video Handling

def clean_title(title: str) -> str:
    """Sanitize video title for safe file naming."""
    title = re.sub(r'[^\w\s-]', '', title)  # Remove special characters
    return re.sub(r'\s+', ' ', title).strip()  # Normalize spaces

def get_youtube_video_links(api_key: str, keyword: str, max_results: int) -> List[Dict[str, str]]:
    """Fetch YouTube video links with rate limit handling."""
    video_links = []
    try:
        youtube = build("youtube", "v3", developerKey=api_key)
        next_page_token = None
        total_fetched = 0

        while total_fetched < max_results:
            try:
                search_request = youtube.search().list(
                    part="snippet",
                    q=keyword,
                    type="video",
                    maxResults=min(MAX_RESULTS_PER_PAGE, max_results - total_fetched),
                    pageToken=next_page_token
                )
                response = search_request.execute()
                for item in response.get("items", []):
                    video_id = item["id"]["videoId"]
                    video_title = item["snippet"]["title"]
                    video_url = f"https://www.youtube.com/watch?v={video_id}"
                    video_links.append({"Title": clean_title(video_title), "Link": video_url})
                    total_fetched += 1
                    if total_fetched >= max_results:
                        break
                next_page_token = response.get("nextPageToken")
                if not next_page_token:
                    break
            except googleapiclient.errors.HttpError as e:
                if e.resp.status == 429:
                    logger.warning("YouTube API rate limit hit. Sleeping for 60s.")
                    time.sleep(60)
                else:
                    raise
    except Exception as e:
        logger.error(f"Failed to fetch video links: {e}")
    return video_links

def save_video_links_to_excel(api_key: str, keyword: str, max_results: int) -> Optional[str]:
    """Save fetched YouTube links to an Excel file."""
    safe_keyword = keyword.replace(" ", "_")
    directory = Path(f"./output_files/{safe_keyword}/processing")
    directory.mkdir(parents=True, exist_ok=True)
    filename = directory / "youtube_video_links.xlsx"

    links = get_youtube_video_links(api_key, keyword, max_results)
    if not links:
        logger.warning(f"No video links retrieved for keyword: {keyword}")
        return None

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Video_Links"
        sheet.append(["Title", "Link"])
        for video in links:
            sheet.append([video["Title"], video["Link"]])
        workbook.save(filename)
        logger.info(f"Saved video links to {filename}")
        return str(filename)
    except Exception as e:
        logger.error(f"Failed to save video links to Excel: {e}")
        return None

def extract_youtube_links(file_path: str, name: str) -> Optional[str]:
    """Extract YouTube links from a file."""
    safe_name = name.replace(" ", "_")
    directory = Path(f"./output_files/{safe_name}/processing")
    directory.mkdir(parents=True, exist_ok=True)
    output_file = directory / "extracted_youtube_links.xlsx"
    youtube_regex = r"(https?://(?:www\.)?youtube\.com/watch\?v=[\w-]+|https?://youtu\.be/[\w-]+)"
    links = []

    try:
        if file_path.lower().endswith(".csv"):
            df = pd.read_csv(file_path)
            for col in df.columns:
                links.extend(re.findall(youtube_regex, str(value)) for value in df[col].astype(str))
        elif file_path.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in wb:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.hyperlink and re.match(youtube_regex, cell.hyperlink.target):
                            links.append(cell.hyperlink.target)
                        elif isinstance(cell.value, str):
                            links.extend(re.findall(youtube_regex, cell.value))
        else:
            logger.error(f"Unsupported file format: {file_path}")
            return None

        links = [link for sublist in links for link in sublist if link]  # Flatten and filter
        if not links:
            logger.warning(f"No YouTube links found in {file_path}")
            return None

        pd.DataFrame({"Youtube_Links": links}).to_excel(output_file, index=False)
        logger.info(f"Extracted links saved to {output_file}")
        return str(output_file)
    except Exception as e:
        logger.error(f"Failed to extract links: {e}")
        return None

def fetch_youtube_video_data(video_url: str) -> Optional[Dict[str, any]]:
    """Fetch metadata for a YouTube video."""
    try:
        ydl_opts = {"quiet": True, "skip_download": True}
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
        return {
            "title": info.get("title", ""),
            "upload_date": info.get("upload_date", ""),
            "views": info.get("view_count", 0),
            "duration": info.get("duration", 0),
            "uploader": info.get("uploader", ""),
            "like_count": info.get("like_count", 0),
            "comment_count": info.get("comment_count", 0)
        }
    except yt_dlp.utils.DownloadError as e:
        logger.error(f"Download error fetching data for {video_url}: {e}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error fetching data for {video_url}: {e}")
        return None

def update_excel_with_video_data(excel_file: str) -> Optional[str]:
    """Update Excel file with YouTube video metadata."""
    try:
        df = pd.read_excel(excel_file)
        columns = ["Title", "Upload Date", "Views", "Video Length (sec)", "Uploaded By",
                   "Like Count", "Comment Count", "File_Name"]
        for col in columns:
            if col not in df.columns:
                df[col] = None

        for index, row in df.iterrows():
            url = row.get("Youtube_Links")
            if pd.notna(url):
                data = fetch_youtube_video_data(url)
                if data:
                    cleaned_title = clean_title(data["title"])
                    file_name = f"{cleaned_title[:30]}.mp4"
                    df.loc[index, columns] = [
                        cleaned_title, data["upload_date"], data["views"], data["duration"],
                        data["uploader"], data["like_count"], data["comment_count"], file_name
                    ]

        df.to_excel(excel_file, index=False)
        logger.info(f"Updated Excel with video data: {excel_file}")
        return excel_file
    except Exception as e:
        logger.error(f"Failed to update Excel: {e}")
        return None

def download_youtube_video(url: str, output_path: str, name: str) -> bool:
    """Download a YouTube video in low quality."""
    file_name = f"{name}.mp4"
    file_path = os.path.join(output_path, file_name)
    if os.path.exists(file_path):
        logger.info(f"{file_name} already exists, skipping download")
        return True

    try:
        ydl_opts = {
            "outtmpl": file_path,
            "format": "worstvideo[ext=mp4]+worstaudio[ext=m4a]/mp4",
            "merge_output_format": "mp4",
            "quiet": True,
            "no_warnings": True
        }
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])
        logger.info(f"Downloaded {file_name}")
        return True
    except yt_dlp.utils.DownloadError as e:
        logger.error(f"Download failed for {url}: {e}")
        return False
    except Exception as e:
        logger.error(f"Unexpected error downloading {url}: {e}")
        return False

def downloader(file_path: str, folder_name: str) -> Optional[str]:
    """Download multiple YouTube videos."""
    safe_folder = folder_name.replace(" ", "_")
    download_path = os.path.join("./videos", safe_folder)
    os.makedirs(download_path, exist_ok=True)

    try:
        df = pd.read_excel(file_path)
        if "Downloaded" not in df.columns:
            df["Downloaded"] = False

        total = len(df)
        success_count = 0
        progress_bar = st.progress(0)
        status_text = st.empty()

        for i, row in df.iterrows():
            url = row.get("Youtube_Links")
            name = (row["File_Name"].replace(".mp4", "").replace(" ", "_")
                    if "File_Name" in df.columns and pd.notna(row["File_Name"])
                    else f"video_{i}")
            if url:
                status_text.text(f"Downloading video {i+1}/{total}")
                success = download_youtube_video(url, download_path, name)
                df.at[i, "Downloaded"] = success
                if success:
                    success_count += 1
            progress_bar.progress((i + 1) / total)

        df.to_excel(file_path, index=False)
        st.write(f"Downloaded {success_count}/{total} videos")
        return download_path if success_count > 0 else None
    except Exception as e:
        logger.error(f"Downloader failed: {e}")
        return None

### AI Analysis

def wait_for_file_active(file) -> Optional[any]:
    """Wait for Generative AI file to become active."""
    try:
        file_status = genai.get_file(file.name)
        while file_status.state.name == "PROCESSING":
            time.sleep(10)
            file_status = genai.get_file(file.name)
        return file_status if file_status.state.name == "ACTIVE" else None
    except Exception as e:
        logger.error(f"Error waiting for file {file.name}: {e}")
        return None

def flatten_json(data: Dict) -> Dict:
    """Flatten nested JSON for Excel compatibility."""
    flat = {}
    for key, value in data.items():
        if isinstance(value, dict):
            for sub_key, sub_value in flatten_json(value).items():
                flat[f"{key}_{sub_key}"] = sub_value
        else:
            flat[key] = value
    return flat

def analyze_video(video_path: str, retries: int = RETRY_ATTEMPTS) -> Optional[Dict]:
    """Analyze a video using Gemini 1.5 Flash with retry logic."""
    for attempt in range(retries):
        try:
            file = genai.upload_file(video_path, mime_type="video/mp4")
            if not file:
                raise ValueError("File upload failed")

            if not wait_for_file_active(file):
                raise ValueError(f"File {file.name} processing failed")

            chat_session = genai.GenerativeModel(
                model_name="gemini-1.5-flash",
                generation_config={
                    "temperature": 1,
                    "top_p": 0.95,
                    "top_k": 64,
                    "max_output_tokens": 8192,
                    "response_mime_type": "application/json"
                }
            ).start_chat(history=[{"role": "user", "parts": [file]}])

            prompt = (
                "For the given video, analyze the video and provide me the following details in JSON format:\n"
                "\"details\": {\n"
                "    \"video_summary\": \"Provide a clear descriptive summary of the video's content.\",\n"
                "    \"number_of_characters\": \"Count the number of distinct characters or people appearing in the video that have impact on the video.\",\n"
                "    \"video_total_emotion\": \"Analyze the overall emotional tone of the video (e.g., positive, negative, neutral).\",\n"
                "    \"video_duration\": \"Provide the duration of the video in seconds.\",\n"
                "    \"video_language\": \"Identify the primary language spoken in the video.\",\n"
                "    \"video_genre\": \"Identify the genre or category of the video (e.g., comedy, drama, action).\",\n"
                "    \"video_mood\": \"Identify the mood or atmosphere of the video (e.g., suspenseful, romantic, humorous).\",\n"
                "    \"video_tone\": \"Identify the tone or style of the video (e.g., serious, light-hearted, satirical).\",\n"
                "    \"video_setting\": \"Identify the primary setting or location of the video.\",\n"
                "    \"video_theme\": \"Identify the central theme or message of the video.\",\n"
                "    \"video_impact\": \"Analyze the impact or significance of the video on the viewer.\",\n"
                "    \"video_target_audience\": \"Identify the target audience or demographic for the video.\",\n"
                "    \"video_influences\": \"Identify any notable influences or inspirations for the video.\",\n"
                "    \"video_references\": \"Identify any references or allusions made in the video.\",\n"
                "    \"color_palette\": \"Identify the primary color palette used in the video.\",\n"
                "    \"words_used\": \"Identify any key words or phrases that are repeated or emphasized in the video.\",\n"
                "    \"brand_name\": \"Identify the brand name associated with the video, especially if it's an advertisement.\",\n"
                "    \"product_name\": \"Identify the product or service being promoted in the video.\",\n"
                "    \"product_features\": \"Identify the key features or benefits of the product or service.\",\n"
                "    \"product_target_audience\": \"Identify the target audience or demographic for the product or service.\",\n"
                "    \"product_message\": \"Identify the central message or value proposition of the product or service.\",\n"
                "    \"product_call_to_action\": \"Identify the call-to-action or desired response from the viewer.\",\n"
                "    \"Personality Trait\": \"Identify the personality of the person in the video. Categories include Extraversion, Agreeableness, Neuroticism, Conscientiousness, Openness to Experience, Narcissism, Machiavellianism, Psychopathy, Maverickism.\",\n"
                "    \"Terms Associated\": \"Identify the terms associated with personality traits\",\n"
                "    \"This person will be\": \"Identify the personality of the person in the video (e.g., 'Will demonstrate energy and assertiveness' or 'Will act whimsically without regard for others').\",\n"
                "    \"Person Tone\": \"Identify the tone of the person in the video (e.g., Serious, Light-hearted, Satirical).\",\n"
                "    \"Person Mood\": \"Identify the mood of the person in the video (e.g., Suspenseful, Romantic, Humorous).\",\n"
                "    \"Person emotion\": \"Identify the emotion of the person in the video (e.g., Positive, Negative, Neutral).\"\n"
                "}"
            )

            response = chat_session.send_message(prompt)
            return json.loads(response.text)
        except Exception as e:
            if attempt < retries - 1 and "rate limit" in str(e).lower():
                sleep_time = SLEEP_TIME_BETWEEN_ANALYSES * (BACKOFF_FACTOR ** attempt)
                logger.warning(f"Rate limit hit, retrying in {sleep_time}s: {e}")
                time.sleep(sleep_time)
            else:
                logger.error(f"Failed to analyze {video_path} after {retries} attempts: {e}")
                return None

def analyze_folder(folder_path: str, project_name: str) -> Optional[str]:
    """Analyze all videos in a folder and save results."""
    safe_project = project_name.replace(" ", "_")
    directory = Path(f"./output_files/{safe_project}/processing")
    directory.mkdir(parents=True, exist_ok=True)
    output_excel = directory / "ai_analyzed_video_data.xlsx"

    try:
        videos = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".mp4")]
        if not videos:
            logger.warning(f"No videos found in {folder_path}")
            return None

        total = len(videos)
        success_count = 0
        all_data = []
        progress_bar = st.progress(0)
        status_text = st.empty()

        for i, video in enumerate(videos):
            status_text.text(f"Analyzing video {i+1}/{total}")
            data = analyze_video(video)
            if data:
                flat_data = flatten_json(data)
                flat_data["File_Name"] = os.path.basename(video)
                all_data.append(flat_data)
                success_count += 1
            progress_bar.progress((i + 1) / total)
            time.sleep(SLEEP_TIME_BETWEEN_ANALYSES)

        if all_data:
            pd.DataFrame(all_data).to_excel(output_excel, index=False)
            st.write(f"Analyzed {success_count}/{total} videos successfully")
            logger.info(f"Analysis saved to {output_excel}")
            return str(output_excel)
        return None
    except Exception as e:
        logger.error(f"Failed to analyze folder {folder_path}: {e}")
        return None

### Data Processing

def merge_data(excel_file1: str, excel_file2: str, project_name: str) -> Optional[str]:
    """Merge two Excel files on File_Name."""
    safe_project = project_name.replace(" ", "_")
    directory = Path(f"./output_files/{safe_project}")
    directory.mkdir(parents=True, exist_ok=True)
    output_file = directory / f"{safe_project}_merged_videos_data.xlsx"

    try:
        df1 = pd.read_excel(excel_file1)
        df2 = pd.read_excel(excel_file2)
        merged = pd.merge(df1, df2, how="inner", on="File_Name")
        merged.to_excel(output_file, index=False)
        logger.info(f"Merged data saved to {output_file}")
        return str(output_file)
    except Exception as e:
        logger.error(f"Failed to merge data: {e}")
        return None

def background_process(keyword: str, max_results: int, folder_name: str) -> Optional[str]:
    """Execute the full pipeline in the background."""
    logger.info(f"Starting background process for keyword: {keyword}")
    steps = [
        ("Fetching video links", lambda: save_video_links_to_excel(YOUTUBE_API_KEY, keyword, max_results)),
        ("Extracting links", lambda x: extract_youtube_links(x, keyword)),
        ("Updating metadata", update_excel_with_video_data),
        ("Downloading videos", lambda x: downloader(x, keyword)),
        ("Analyzing videos", lambda x: analyze_folder(x, keyword)),
        ("Merging data", lambda x, y: merge_data(x, y, keyword))
    ]

    result = None
    excel_file = None
    summary_file = None

    for step_name, func in steps:
        logger.info(f"Executing: {step_name}")
        if step_name == "Merging data":
            result = func(excel_file, summary_file) if excel_file and summary_file else None
        elif step_name in ["Fetching video links", "Updating metadata"]:
            result = func() if step_name == "Fetching video links" else func(result)
            excel_file = result if step_name == "Updating metadata" else excel_file
        elif step_name == "Analyzing videos":
            result = func(result)
            summary_file = result
        else:
            result = func(result)
        if not result:
            logger.error(f"{step_name} failed, aborting process")
            return None

    if result:
        upload_videos_to_s3(downloader(excel_file, keyword), folder_name)
        upload_final_outputs_to_s3(result)
        logger.info("Background process completed successfully")
    return result

### Streamlit App

def list_output_files() -> List[str]:
    """List files in the S3 output bucket."""
    s3 = boto3.client("s3")
    try:
        response = s3.list_objects_v2(Bucket=S3_OUTPUT_BUCKET)
        return [obj["Key"] for obj in response.get("Contents", [])]
    except ClientError as e:
        logger.error(f"Failed to list S3 output files: {e}")
        return []

def app():
    """Main Streamlit application."""
    if not all([YOUTUBE_API_KEY, GENAI_API_KEY, S3_BUCKET]):
        st.error("Missing required environment variables (YOUTUBE_API_KEY, GENAI_API_KEY, S3_BUCKET)")
        return

    genai.configure(api_key=GENAI_API_KEY)
    st.sidebar.title("YouTube Video Summarizer")
    option = st.sidebar.radio("Mode", ["Keyword Search", "Single Video", "File Mode",
                                       "Background Tasks Status", "Show Output Files"])

    if option == "Keyword Search":
        st.title("Keyword Search (Background Processing)")
        keyword = st.text_input("Keyword")
        max_results = st.slider("Number of Videos", 1, 50, 10)
        if st.button("Start"):
            if keyword:
                task_id = run_background_task(background_process, keyword, max_results,
                                             task_keyword=keyword, folder_name=keyword)
                st.success(f"Task started (ID: {task_id}). Check status in 'Background Tasks Status'.")
            else:
                st.warning("Please enter a keyword")

    elif option == "Single Video":
        st.title("Single Video Analysis")
        url = st.text_input("YouTube URL")
        if st.button("Analyze"):
            if url:
                try:
                    data = fetch_youtube_video_data(url)
                    if not data:
                        st.error("Failed to fetch video data")
                        return
                    file_name = clean_title(data["title"])[:30].replace(" ", "_")
                    video_path = "./videos/single_url"
                    os.makedirs(video_path, exist_ok=True)
                    if download_youtube_video(url, video_path, file_name):
                        full_path = os.path.join(video_path, f"{file_name}.mp4")
                        if os.path.exists(full_path):
                            output = analyze_video(full_path)
                            if output:
                                st.json(output)
                                upload_videos_to_s3(video_path)
                            else:
                                st.error("Analysis failed")
                        else:
                            st.error("Video file not found after download")
                    else:
                        st.error("Download failed")
                except Exception as e:
                    st.error(f"Error: {e}")
            else:
                st.warning("Please enter a URL")

    elif option == "File Mode":
        st.title("File Mode")
        file_path = st.text_input("File Path (CSV/Excel)")
        project_name = st.text_input("Project Name")
        if st.button("Process"):
            if file_path and project_name:
                result = extract_youtube_links(file_path, project_name)
                if result:
                    result = update_excel_with_video_data(result)
                    if result:
                        video_folder = downloader(result, project_name)
                        if video_folder:
                            summary = analyze_folder(video_folder, project_name)
                            if summary:
                                final_file = merge_data(result, summary, project_name)
                                if final_file:
                                    upload_videos_to_s3(video_folder)
                                    upload_final_outputs_to_s3(final_file)
                                    st.success("Processing complete")
                                else:
                                    st.error("Merging failed")
                            else:
                                st.error("Analysis failed")
                        else:
                            st.error("Download failed")
                    else:
                        st.error("Metadata update failed")
                else:
                    st.error("Link extraction failed")
            else:
                st.warning("Please provide file path and project name")

    elif option == "Background Tasks Status":
        st.title("Background Tasks Status")
        tasks = load_background_tasks()
        if tasks:
            for task in tasks.values():
                st.json(task)
        else:
            st.info("No tasks found")

    elif option == "Show Output Files":
        st.title("Output Files")
        files = list_output_files()
        if files:
            for file in files:
                st.markdown(f"- {file} - [Download](https://{S3_OUTPUT_BUCKET}.s3.amazonaws.com/{file})")
        else:
            st.info("No files found")

if __name__ == "__main__":
    app()